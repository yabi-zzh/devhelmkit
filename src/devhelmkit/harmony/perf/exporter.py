# !/usr/bin/env python
# -*- coding: utf-8 -*-
"""性能数据 Excel 导出。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from devhelmkit.harmony.perf.models import PerfDataPoint

_THIN_GRAY = Side(style="thin", color="D1D5DB")
_THIN_LIGHT = Side(style="thin", color="E5E7EB")

_GROUP_COLORS = {
    "time": "FF6B7280",
    "fps": "FF3B82F6",
    "cpu": "FF6366F1",
    "gpu": "FFEC4899",
    "memory": "FF10B981",
    "network": "FFF59E0B",
}

_GROUP_LABELS = {
    "time": "时间",
    "fps": "FPS",
    "cpu": "CPU",
    "gpu": "GPU",
    "memory": "内存",
    "network": "网络",
}


def _to_num(value: Optional[float], integer: bool = False):
    if value is None:
        return None
    try:
        if value != value:  # NaN
            return None
    except TypeError:
        return None
    if integer:
        return int(round(value))
    return round(value * 100) / 100


def _lighten(argb: str, factor: float = 0.45) -> str:
    r = int(argb[2:4], 16)
    g = int(argb[4:6], 16)
    b = int(argb[6:8], 16)
    r = min(255, r + round((255 - r) * factor))
    g = min(255, g + round((255 - g) * factor))
    b = min(255, b + round((255 - b) * factor))
    return "FF%02X%02X%02X" % (r, g, b)


def default_perf_export_path(package_name: str, directory: str = ".") -> str:
    """生成默认导出文件名：perf_<pkg>_<ts>.xlsx。"""
    safe_pkg = (package_name or "data").replace("/", "_").replace("\\", "_")
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return str(Path(directory) / ("perf_%s_%s.xlsx" % (safe_pkg, ts)))


def export_perf_xlsx(
    points: Sequence[PerfDataPoint],
    path: str,
    package_name: str = "",
) -> str:
    """将性能采样点导出为 xlsx。

    Args:
        points: 采样点列表
        path: 输出路径（.xlsx）
        package_name: 写入工作簿标题，不影响表结构

    Returns:
        实际写入的绝对路径
    """
    if not points:
        raise ValueError("无数据可导出")

    core_count = max((len(p.cpu.cores) for p in points), default=0)

    cols = [
        {"header": "Time", "key": "time", "width": 14, "group": "time"},
        {"header": "FPS", "key": "fps", "width": 8, "group": "fps", "integer": True},
        {"header": "RefreshRate", "key": "refresh", "width": 12,
         "group": "fps", "integer": True},
        {"header": "Jank", "key": "jank", "width": 8, "group": "fps", "integer": True},
        {"header": "App %", "key": "cpuApp", "width": 10, "group": "cpu"},
        {"header": "Total %", "key": "cpuTotal", "width": 10, "group": "cpu"},
    ]
    for i in range(core_count):
        cols.append({
            "header": "C%d %%" % i, "key": "c%d" % i, "width": 8, "group": "cpu",
        })
    cols.extend([
        {"header": "GPU %", "key": "gpu", "width": 8, "group": "gpu"},
        {"header": "PSS (MB)", "key": "pss", "width": 12, "group": "memory"},
        {"header": "Heap (MB)", "key": "heap", "width": 12, "group": "memory"},
        {"header": "Down (KB/s)", "key": "netDown", "width": 14, "group": "network"},
        {"header": "Up (KB/s)", "key": "netUp", "width": 14, "group": "network"},
    ])

    wb = Workbook()
    ws = wb.active
    ws.title = "性能数据"
    ws.freeze_panes = "A3"

    # 第 1 行：分组标题；第 2 行：列名
    group_spans: List[dict] = []
    prev_group = ""
    span_start = 1
    for i, col in enumerate(cols):
        group = col["group"]
        if group != prev_group:
            if prev_group:
                group_spans.append({
                    "start": span_start, "end": i, "group": prev_group,
                })
            span_start = i + 1
            prev_group = group
    group_spans.append({
        "start": span_start, "end": len(cols), "group": prev_group,
    })
    boundary_cols = {span["end"] for span in group_spans}

    center = Alignment(horizontal="center", vertical="center")

    for span in group_spans:
        start, end = span["start"], span["end"]
        if start != end:
            ws.merge_cells(
                start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start)
        cell.value = _GROUP_LABELS.get(span["group"], span["group"].upper())
        bg = _GROUP_COLORS.get(span["group"], "FF6B7280")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = center
    ws.row_dimensions[1].height = 24

    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=i, value=col["header"])
        bg = _GROUP_COLORS.get(col["group"], "FF6B7280")
        cell.fill = PatternFill("solid", fgColor=_lighten(bg))
        cell.font = Font(bold=True, color="374151", size=10)
        cell.alignment = center
        cell.border = Border(
            bottom=Side(style="thin", color=bg),
            right=_THIN_GRAY if i in boundary_cols else None,
        )
        ws.column_dimensions[get_column_letter(i)].width = col["width"]
    ws.row_dimensions[2].height = 22

    stripe = PatternFill("solid", fgColor="EEF0F4")
    for row_idx, point in enumerate(points):
        values = [
            point.time_label,
            _to_num(point.fps.fps, integer=True),
            _to_num(point.fps.refresh_rate, integer=True),
            _to_num(
                float(point.fps.jank_count)
                if point.fps.jank_count is not None else None,
                integer=True,
            ),
            _to_num(point.cpu.proc_usage),
            _to_num(point.cpu.total_usage),
        ]
        for i in range(core_count):
            usage = point.cpu.cores[i].usage if i < len(point.cpu.cores) else None
            values.append(_to_num(usage))
        values.extend([
            _to_num(point.gpu.load),
            _to_num(point.memory.pss),
            _to_num(point.memory.heap_size),
            _to_num(point.network.down),
            _to_num(point.network.up),
        ])

        excel_row = row_idx + 3
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = center
            cell.font = Font(size=10)
            if isinstance(value, (int, float)):
                cell.number_format = (
                    "0" if cols[col_idx - 1].get("integer") else "0.00")
            if row_idx % 2 == 0:
                cell.fill = stripe
            if col_idx in boundary_cols:
                cell.border = Border(right=_THIN_LIGHT)

    out = Path(path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.properties.title = package_name or "性能数据"
    wb.save(str(out))
    return str(out)
